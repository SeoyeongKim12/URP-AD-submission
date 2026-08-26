"""
Aim 2 · B1 최종결과 통합 엑셀 (Drive 업로드용) (sjlee)
======================================================
결과요약·4조건·성능·B1 계수/p-value·Gn 절단별 계수를 시트별로 정리한 단일 xlsx 생성.

산출: aim2/aim2_sjlee/B1_최종결과_정리.xlsx  (→ Drive AIM 2/)
입력: ~/Downloads/ { adl_wide.csv, baseline_sample.csv }
"""
from pathlib import Path
import numpy as np
import pandas as pd
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LogisticRegression
from statsmodels.miscmodels.ordinal_model import OrderedModel
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

DOWNLOADS = Path.home() / "Downloads"
OUT = Path(__file__).parent / "aim2_sjlee"; OUT.mkdir(exist_ok=True)
XLSX = OUT / "B1_최종결과_정리.xlsx"
BADL = ["ADL0101", "ADL0102", "ADL0103", "ADL0104", "ADL0105", "ADL0106B"]
IADL = ["ADL0106A", "ADL0107A", "ADL0110A", "ADL0111A", "ADL0112A", "ADL0113A",
        "ADL0114A", "ADL0115A", "ADL0116A", "ADL0116B", "ADL0117A", "Q18",
        "ADL0121A", "ADL0122Q", "ADL0123L"]
ITEMS = BADL + IADL
LAB = {"ADL0101": "Q1 먹기", "ADL0102": "Q2 걷기", "ADL0103": "Q3 화장실", "ADL0104": "Q4 목욕",
       "ADL0105": "Q5 몸단장", "ADL0106B": "Q6b 옷입기", "ADL0106A": "Q6a 옷고르기",
       "ADL0107A": "Q7 전화", "ADL0110A": "Q10 설거지", "ADL0111A": "Q11 식사준비",
       "ADL0112A": "Q12 집안일", "ADL0113A": "Q13 빨래", "ADL0114A": "Q14 가전",
       "ADL0115A": "Q15 외출", "ADL0116A": "Q16a 쇼핑", "ADL0116B": "Q16b 지불",
       "ADL0117A": "Q17 금전", "Q18": "Q18 혼자있기", "ADL0121A": "Q21 글쓰기",
       "ADL0122Q": "Q22 취미", "ADL0123L": "Q23 가전사용"}


def build():
    adl = pd.read_csv(DOWNLOADS / "adl_wide.csv", low_memory=False)
    b = adl[adl["VISITNUM"] == 2.0].copy()
    def col(c): return b[c + "__resolved"] if c + "__resolved" in b.columns else b[c]
    data = {c: ((col("ADL0118A") + col("ADL0118B") + col("ADL0118C")) if c == "Q18" else col(c)) for c in ITEMS}
    X = pd.DataFrame(data); X["STUDYID"] = b["STUDYID"].values; X["USUBJID"] = b["USUBJID"].values
    X = X.drop_duplicates(subset=["STUDYID", "USUBJID"])
    bs = pd.read_csv(DOWNLOADS / "baseline_sample.csv")
    cs = bs[bs["in_common_comparison_sample"] == True][["STUDYID", "USUBJID", "ds_stage"]]
    m = cs.merge(X, on=["STUDYID", "USUBJID"], how="left")
    imp = SimpleImputer(strategy="median").fit(m[ITEMS])
    return pd.DataFrame(imp.transform(m[ITEMS]), columns=ITEMS), m["ds_stage"].astype(int).values


def sheets():
    Xi, y = build()

    # 요약 시트들
    summary = pd.DataFrame({
        "항목": ["연구", "분석일", "표본", "문항(X)", "타깃(y)", "CV", "핵심 결론"],
        "내용": ["Aim2 B1 대안 채점법 — ADCS-ADL로 실측 DS 단계 예측",
                "2026-08-05", "공통표본 2,203 (AD-1061/1063/1064)",
                "A0 입력과 동일 21문항(iADL15+bADL6, __resolved)",
                "실측 DS 단계 ds_stage (0-5)",
                "leave-one-trial-out 3-fold (누출 0, 시험변수 미투입)",
                "원 B1은 중증 과소분류로 미충족 → Gn(부분비례오즈 엘라스틱넷+비대칭임계+중첩CV마진)이 4조건 전부 충족(주 0.10)"]})

    cond = pd.DataFrame({
        "조건": ["1. MAE 감소 ≥0.10", "2. 2↑과소 증가 ≤2%p", "3. 4·5→≤3 증가 ≤2%p", "4. ≥2시험 MAE감소", "종합 판정"],
        "기준": ["≥0.10", "≤+2%p", "≤+2%p", "≥2/3", "4조건"],
        "원 B1": ["+0.172 ✅", "+0.8%p ✅", "+35.2%p ❌", "3/3 ✅", "미충족"],
        "Gn(최종)": ["+0.126 ✅", "-0.5%p ✅", "-2.3%p ✅(여유 +4.3%p)", "3/3 ✅", "★ 4조건 충족"]})

    perf = pd.DataFrame({
        "지표": ["MAE", "MAE 감소(vs A1)", "가중카파", "실제 4·5→≤3", "2↑단계 과소평가"],
        "원 B1": ["0.419", "+0.172", "0.570", "62.0%", "3.5%"],
        "Gn(최종)": ["0.465", "+0.126", "0.601", "24.5%", "2.2%"],
        "A1(2015)": ["0.591", "—", "0.498", "26.8%", "2.7%"]})

    # B1 비례오즈 계수·p (무벌점)
    res = OrderedModel(y, Xi.values, distr="logit").fit(method="bfgs", disp=False, maxiter=300)
    po = pd.DataFrame({"문항": [LAB[c] for c in ITEMS], "코드": ITEMS,
                       "계수(β)": res.params[:len(ITEMS)], "OR": np.exp(res.params[:len(ITEMS)]),
                       "SE": res.bse[:len(ITEMS)], "z": res.tvalues[:len(ITEMS)],
                       "p_value": res.pvalues[:len(ITEMS)]})
    po["유의"] = po["p_value"].apply(lambda p: "***" if p < .001 else "**" if p < .01 else "*" if p < .05 else "")
    po = po.sort_values("p_value").reset_index(drop=True)
    po["계수(β)"] = po["계수(β)"].round(3); po["OR"] = po["OR"].round(2)
    po["SE"] = po["SE"].round(3); po["z"] = po["z"].round(2); po["p_value"] = po["p_value"].map(lambda x: f"{x:.2e}")

    # Gn 엘라스틱넷 절단별 계수 (원 눈금)
    sc = StandardScaler().fit(Xi.values); Z = sc.transform(Xi.values); sd = sc.scale_
    cut = {}
    for k in range(1, 6):
        lr = LogisticRegression(penalty="elasticnet", solver="saga", l1_ratio=0.5, C=0.1,
                                max_iter=6000, tol=1e-3).fit(Z, (y >= k).astype(int))
        cut[k] = lr.coef_[0] / sd
    gn = pd.DataFrame({"문항": [LAB[c] for c in ITEMS], "코드": ITEMS})
    for k in range(1, 6):
        gn[f"P(Y≥{k})"] = [round(cut[k][i], 2) if abs(cut[k][i]) > 1e-4 else 0 for i in range(len(ITEMS))]

    notes = pd.DataFrame({"한계·주의": [
        "1. p-value는 무벌점 B1(비례오즈)에서만 타당 — 벌점(Gn)은 표준오차 편향으로 p 무효.",
        "2. Gn 계수는 표준화 후 적합→원 눈금 환산. 음수=문항점수 높을수록(독립적) 저단계.",
        "3. Gn은 앙상블(비평행 로짓+임계)이라 단일 채점표 취지 약화 — 해석은 B1 계수, 성능은 Gn.",
        "4. 3시험 소표본 CV라 외부검증 전 성능은 낙관적.",
        "5. 상위 절단 P(Y≥4·5)=기본 ADL 주도, 하위 절단 P(Y≥1·2)=iADL 주도 → 중증 검출 메커니즘.",
        "6. 모든 튜닝(λ·마진·τ) fold/inner 훈련 내부, 시험변수 미투입(누출 0)."]})

    return {"요약": summary, "추가가치_4조건": cond, "성능지표": perf,
            "B1_계수_pvalue": po, "Gn_절단별계수": gn, "한계_주의": notes}


def style(path, data):
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        for name, df in data.items():
            df.to_excel(w, sheet_name=name, index=False)
        wb = w.book
        hdr_fill = PatternFill("solid", fgColor="1F4E78"); hdr_font = Font(bold=True, color="FFFFFF")
        for name, df in data.items():
            ws = wb[name]
            for j, colname in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=j); cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                width = max(10, min(48, int(df[colname].astype(str).map(len).max() if len(df) else 10) + 3,
                                    max(len(str(colname)) + 3, 12)))
                ws.column_dimensions[get_column_letter(j)].width = max(12, len(str(colname)) + 4)
            ws.freeze_panes = "A2"
    print(f">>> 저장: {path}")


if __name__ == "__main__":
    style(XLSX, sheets())
